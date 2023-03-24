import tkinter as tk
from datetime import datetime
from openpyxl import Workbook, load_workbook
import datetime
from datetime import datetime, timedelta
import requests
from PIL import Image, ImageTk
import threading
import time
from forex_python.converter import CurrencyRates
import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
import pandas as pd
from PIL import Image, ImageTk
import PIL.Image
import PIL.ImageTk
#создание нового файла Excel или загрузка существующего
try:
    wb = load_workbook("Управление финансами.xlsx")
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"
    ws.append(["ID", "Категория", "Сумма", "Дата"])
else:
    ws = wb.active
#добавление нового расхода
def set_budget():
    budget = float(budget_entry.get() or 0)
    budget_label.config(text=f"Ежемесячный бюджет: {budget:.2f}")
     # Add budget data to cell E1
    budget_data = budget_entry.get()
    sheet = wb.active
    sheet['E1'] = 'Ежемесячный бюджет'
    sheet['E2'] = budget_data
    # Save the Excel file
    wb.save("Управление финансами.xlsx")
def analyze_expenses_by_category():
    # Загружаем данные из файла Excel
    df = pd.read_excel("Управление финансами.xlsx", sheet_name="Expenses")
    # Группируем расходы по категориям и суммируем их
    expenses_by_category = df.groupby("Категория")["Сумма"].sum()
    # Визуализируем расходы по категориям
    plt.figure(figsize=(8, 6))
    plt.pie(expenses_by_category, labels=expenses_by_category.index, autopct='%1.1f%%')
    plt.title("Расходы по категориям")
    plt.show()
def add_expense(category, amount, date):
    #добавление данных в файл Excel
    id = len(ws["A"])  # ID - это номер строки
    ws.append([id, category, amount, date])
    wb.save("Управление финансами.xlsx")
#отображение списка расходов
def list_expenses():
    for row in ws.iter_rows(min_row=2, values_only=True):
        print(row)
#расчет налогов и вычетов
def predict_future_expenses():
    #Загружаем данные из файла Excel
    df = pd.read_excel("Управление финансами.xlsx", sheet_name="Expenses")
    #Определяем текущие расходы
    current_expenses = df["Сумма"].sum()
    #Вычисляем скользящее среднее за последние 3 месяца
    rolling_mean = df["Сумма"].rolling(window=3).mean().iloc[-1]
    #Прогнозируем будущие расходы на основе скользящего среднего
    future_expenses = rolling_mean * 1.1  #увеличиваем на 10%
    return future_expenses, current_expenses
def predict_expenses():
    future_expenses, current_expenses = predict_future_expenses()
    expenses_result.config(text=f"Текущие расходы: {current_expenses}\nПрогноз расходов на следующий месяц: {future_expenses:.2f}")
def calculate_tax():
    income = float(income_entry.get() or 0)
    expenses = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        expenses.append(row[2])
    total_expenses = sum(expenses)
    deduction = min(total_expenses, 40000)
    taxable_income = max(income - deduction, 0)
    tax_rate = 0.13  #ставка налога на доходы физических лиц
    tax = taxable_income * tax_rate
    after_tax_income = income - tax
    tax_result.config(text=f"Доход: {income}\nРасходы: {total_expenses}\nВычет: {deduction}\nНалогооблагаемый доход: {taxable_income}\nСтавка налога: {tax_rate:.0%}\nНалог: {tax:.2f}\nПосле уплаты налога: {after_tax_income:.2f}")
    #Рисование диаграммы расходов
    categories = []
    expenses = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        categories.append(row[1])
        expenses.append(row[2])
    fig, ax = plt.subplots()
    ax.bar(categories, expenses)
    ax.set_title("Расходы по категориям")
    ax.set_xlabel("Категории")
    ax.set_ylabel("Сумма, руб.")
    #Вставка диаграммы в окно приложения
    canvas = FigureCanvasTkAgg(fig, master=root)
    canvas.draw()
    canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
    #Создание панели инструментов для диаграммы
    toolbar = NavigationToolbar2Tk(canvas, root)
    toolbar.update()
    canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
#создание главного окна
def get_currency_rates(api_key):
    url = f"https://api.exchangerate-api.com/v4/latest/USD"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}"
    }
    response = requests.get(url, headers=headers)
    if response.status_code != 200:
        print("Ошибка при получении курсов валют")
        return
#bb857173939c485304291549 токен для ввода
    currency_rates = response.json()
    usd_rate = currency_rates["rates"]["RUB"]
    eur_rate = currency_rates["rates"]["EUR"] / currency_rates["rates"]["USD"]
    eur_to_rub_rate = currency_rates["rates"]["RUB"] / eur_rate
    usd_rate_label.config(text=f"1 USD = {usd_rate:.2f} RUB")
    eur_to_rub_rate_label.config(text=f"1 EUR = {eur_to_rub_rate:.2f} RUB")
root = tk.Tk()
loading_gif = PIL.Image.open("zag1.png")
#задаем размеры окна
root.geometry("800x800")
root.attributes("-fullscreen", True)
label1 = tk.Label(root, text="Управление личными финансами",font=("Playfair Display", 12))
# Создать объект PhotoImage
loading_gif_image = PIL.ImageTk.PhotoImage(loading_gif)
loading_label = tk.Label(root, image=loading_gif_image)
loading_label.place(x=0, y=0, relwidth=1, relheight=1)
root.update()
# Задержка в несколько секунд
time.sleep(5)
# Удалить гифку и отобразить основную программу
loading_label.destroy()
image = Image.open("background.png")
background_image = ImageTk.PhotoImage(image)
root.attributes("-fullscreen", True)
#установка изображения в качестве фона
background_label = tk.Label(root, image=background_image)
background_label.place(x=0, y=0, relwidth=1, relheight=1)
#задаем заголовок окна
root.title("Управление личными финансами")
#задаем размеры окна
root.geometry("800x800")
#создание виджетов
label = tk.Label(root, text="Управление личными финансами",font=("Playfair Display", 12))
label.config(borderwidth=2, foreground="black", bg="#f2c7aa", highlightbackground="red")

#создание поля ввода для дохода
income_label = tk.Label(root, text="Доход",font=("Playfair Display", 12))
income_label.config(borderwidth=2, foreground="black", bg="#fafaaa", highlightbackground="red")

income_entry = tk.Entry(root,font=("Playfair Display", 12))
income_entry.config(borderwidth=2, foreground="black", bg="white", highlightbackground="red")
#создание поля ввода для категории
category_label = tk.Label(root, text="Категория",font=("Playfair Display", 12))
category_label.config(borderwidth=2, foreground="black", bg="#fafaaa", highlightbackground="red")
category_entry = tk.Entry(root,font=("Playfair Display", 12))
category_entry.config(borderwidth=2, foreground="black", bg="white", highlightbackground="red")
#создание поля ввода для суммы
amount_label = tk.Label(root, text="Сумма",font=("Playfair Display", 12))
amount_label.config(borderwidth=2, foreground="black", bg="#fafaaa", highlightbackground="red")
amount_entry = tk.Entry(root,font=("Playfair Display", 12))
amount_entry.config(borderwidth=2, foreground="black", bg="white", highlightbackground="red")
#создание поля ввода для даты
date_label = tk.Label(root, text="Дата",font=("Playfair Display", 12))
date_label.config(borderwidth=2, foreground="black", bg="#fafaaa", highlightbackground="red")
date_entry = tk.Entry(root,font=("Playfair Display", 12))
date_entry.config(borderwidth=2, foreground="black", bg="white", highlightbackground="red")
#создание кнопки для добавления расхода
def add_expense_callback():
    amount_str = amount_entry.get()
    if not amount_str:
        return
    add_expense(category_entry.get(), float(amount_str), date_entry.get())
def blink(button, color):
    button.config(bg=color, activebackground=color)
    button.after(500, blink, button, "grey" if button.cget("bg") == "yellow" else "yellow")

button_add = tk.Button(root, text="Добавить расход", command=add_expense_callback, font=("Playfair Display", 12), bg="grey", activebackground="orange", activeforeground="white", highlightthickness=0, bd=2, relief="groove")
button_add.after(500, blink, button_add, "yellow")
#создание поля для отображения результатов расчета налогов и вычетов
tax_result = tk.Label(root,text="", font=("Playfair Display", 12), bg="white", bd=2, relief="groove", highlightthickness=2, highlightbackground="gray", highlightcolor="white")

tax_result.config(borderwidth=2, foreground="black", bg="beige", highlightbackground="red")#config для изменения цвета label и кнопки

button_tax = tk.Button(root, text="Рассчитать налоги и вычеты", command=calculate_tax, font=("Playfair Display", 12), bd=2, relief="groove", highlightthickness=2, highlightbackground="gray", highlightcolor="white", activebackground="orange", activeforeground="white")
button_tax.config(borderwidth=2, foreground="black", bg="#fafaaa", highlightbackground="red")

api_key_label = tk.Label(root, text="Введите API ключ:",font=("Playfair Display", 12), bd=2, relief="groove", highlightthickness=2, highlightbackground="blue", highlightcolor="white", activebackground="orange", activeforeground="white")
api_key_label.config(borderwidth=2, foreground="black", bg="white", highlightbackground="red")

api_key_entry = tk.Entry(root,text="bb857173939c485304291549",font=("Playfair Display", 12))#при необходимости удалить text="bb857173939c485304291549"
api_key_entry.config(borderwidth=2, foreground="black", bg="white", highlightbackground="red")

usd_rate_label = tk.Label(root, text="1 USD = ? RUB",font=("Playfair Display", 10), bd=2, relief="groove", highlightthickness=2, highlightbackground="gray", highlightcolor="white", activebackground="orange", activeforeground="white")
usd_rate_label.config(borderwidth=2, foreground="black", bg="white", highlightbackground="red")

eur_to_rub_rate_label = tk.Label(root, text="1 EUR = ? RUB",font=("Playfair Display", 10), bd=2, relief="groove", highlightthickness=2, highlightbackground="gray", highlightcolor="white", activebackground="orange", activeforeground="white")
eur_to_rub_rate_label.config(borderwidth=2, foreground="black", bg="white", highlightbackground="red")

button_currency = tk.Button(root, text="Получить курс валют", command=lambda: get_currency_rates(api_key_entry.get()),font=("Playfair Display", 12), bd=2, relief="groove", highlightthickness=2, highlightbackground="gray", highlightcolor="white", activebackground="orange", activeforeground="white")
button_currency.config(borderwidth=2, foreground="black", bg="#fafaaa", highlightbackground="red")#t
#создание кнопки для прогнозирования расходов
button_predict = tk.Button(root, text="Прогнозировать расходы", command=predict_expenses,font=("Playfair Display", 12), bd=2, relief="groove", highlightthickness=2, highlightbackground="gray", highlightcolor="white", activebackground="orange", activeforeground="white")
button_predict.config(borderwidth=2, foreground="black", bg="#fafaaa", highlightbackground="red")
#создание поля для отображения результатов
expenses_result = tk.Label(root, text="",font=("Playfair Display", 12), bd=2, relief="groove", highlightthickness=2, highlightbackground="gray", highlightcolor="white", activebackground="orange", activeforeground="white")
expenses_result.config(borderwidth=2, foreground="black", bg="white", highlightbackground="red")

button_analyze_by_category = tk.Button(root, text="Анализировать расходы по категориям", command=analyze_expenses_by_category,font=("Playfair Display", 12))
button_analyze_by_category.config(borderwidth=2, foreground="black", bg="#fafaaa", highlightbackground="red")
#размещение виджетов на окне
budget_label = tk.Label(root, text="Ежемесячный бюджет",bg="red",font=("Playfair Display", 12), bd=2, relief="groove", highlightthickness=2, highlightbackground="gray", highlightcolor="white", activebackground="orange", activeforeground="white")
budget_label.config(borderwidth=2, foreground="black", bg="white", highlightbackground="red")

budget_entry = tk.Entry(root)
budget_entry.config(borderwidth=2, foreground="black", bg="white", highlightbackground="red")

budget_button = tk.Button(root, text="Установить бюджет", command=set_budget,bg="blue",font=("Playfair Display", 12), bd=2, relief="groove", highlightthickness=2, highlightbackground="gray", highlightcolor="white", activebackground="orange", activeforeground="white")
budget_button.config(borderwidth=2, foreground="black", bg="#fafaaa", highlightbackground="red")
#виджеты
label.pack()
income_label.pack()
income_entry.pack()
category_label.pack()
category_entry.pack()
amount_label.pack()
amount_entry.pack()
date_label.pack()
date_entry.pack()
button_add.pack()
button_tax.pack()
tax_result.pack()
expenses_result = tk.Label(root, text="",font=("Playfair Display", 10), bd=2, relief="groove", highlightthickness=2, highlightbackground="gray", highlightcolor="white", activebackground="orange", activeforeground="white")
expenses_result.pack()
api_key_label.pack()
api_key_entry.pack()
button_currency.pack()
usd_rate_label.pack()
eur_to_rub_rate_label.pack()
label.pack()
button_predict.pack()
expenses_result.pack()
button_analyze_by_category.pack()
budget_label.pack()
budget_entry.pack()
budget_button.pack()
#запуск главного цикла приложения
root.mainloop()